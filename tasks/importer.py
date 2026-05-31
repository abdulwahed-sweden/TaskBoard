"""CSV/XLSX task import: parsing, column mapping, and per-row validation.

Everything here is pure and side-effect-free until :func:`commit`, so the parse
→ map → preview pipeline is easy to test and safe to re-run. Validation reuses
the project-type validators (custom fields and status), with a string-coercion
step in front because spreadsheet cells arrive as text.
"""

import csv
import datetime
import io

from django.core.exceptions import ValidationError

from organizations.custom_fields import validate_custom_fields
from organizations.models import FieldDefinition
from organizations.workflow import status_definitions_for, validate_status

from .activity import log_created
from .models import Task

# Core task fields the importer can map: (target, label, required).
CORE_FIELDS = [
    ("title", "Title", True),
    ("notes", "Notes", False),
    ("due_date", "Due date", False),
]
CUSTOM_PREFIX = "custom_"

# Guard against pathologically large uploads; exceeding it is reported, never
# silently truncated.
MAX_ROWS = 1000


class ImportError(ValueError):
    """A problem with the uploaded file itself (not a per-row issue)."""


# --- parsing -----------------------------------------------------------------

def parse_upload(file, filename):
    """Return ``(headers, rows)`` from an uploaded CSV or XLSX file.

    ``rows`` is a list of ``{header: value}`` dicts with string values. Raises
    :class:`ImportError` for an unsupported type, an empty file, or too many
    rows.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        headers, rows = _parse_csv(file)
    elif name.endswith(".xlsx"):
        headers, rows = _parse_xlsx(file)
    else:
        raise ImportError("Unsupported file type — upload a .csv or .xlsx file.")

    if not headers:
        raise ImportError("The file has no header row.")
    if not rows:
        raise ImportError("The file has no data rows.")
    if len(rows) > MAX_ROWS:
        raise ImportError(
            f"Too many rows ({len(rows)}); the import limit is {MAX_ROWS}."
        )
    return headers, rows


def _parse_csv(file):
    data = file.read()
    if isinstance(data, bytes):
        data = data.decode("utf-8-sig")  # tolerate a BOM
    reader = csv.DictReader(io.StringIO(data))
    headers = reader.fieldnames or []
    rows = [
        {h: (row.get(h) or "").strip() for h in headers}
        for row in reader
    ]
    return headers, rows


def _parse_xlsx(file):
    # Imported lazily so the dependency is only needed when XLSX is used.
    from openpyxl import load_workbook

    workbook = load_workbook(file, read_only=True, data_only=True)
    sheet = workbook.active
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows = []
    for values in rows_iter:
        if values is None or all(v is None for v in values):
            continue  # skip blank rows
        row = {}
        for index, header in enumerate(headers):
            value = values[index] if index < len(values) else None
            row[header] = "" if value is None else str(value).strip()
        rows.append(row)
    workbook.close()
    return headers, rows


# --- target fields & mapping -------------------------------------------------

def target_fields(project):
    """Mappable target fields for ``project``: core, status (if the type has a
    workflow), and one ``custom_<name>`` per field definition."""
    fields = list(CORE_FIELDS)
    if status_definitions_for(project):
        fields.append(("status", "Status", False))
    if project.project_type_id is not None:
        for definition in project.project_type.field_definitions.all():
            fields.append(
                (
                    f"{CUSTOM_PREFIX}{definition.name}",
                    definition.label,
                    definition.required,
                )
            )
    return fields


def auto_mapping(project, headers):
    """Best-guess mapping target_field -> source column by case-insensitive
    match on the column header against the target name or label."""
    lookup = {h.lower().strip(): h for h in headers}
    mapping = {}
    for name, label, _required in target_fields(project):
        for key in (name, name.removeprefix(CUSTOM_PREFIX), label):
            match = lookup.get(key.lower())
            if match:
                mapping[name] = match
                break
    return mapping


def map_row(row, mapping):
    """Pull the mapped source values for one row into ``{target_field: value}``,
    dropping unmapped targets and blank values."""
    mapped = {}
    for target, source in mapping.items():
        if not source:
            continue
        value = (row.get(source) or "").strip()
        if value != "":
            mapped[target] = value
    return mapped


# --- per-row validation ------------------------------------------------------

def _coerce_custom(definition, raw):
    """Coerce a string cell to the field's Python type. Raises ``ValueError``."""
    types = FieldDefinition.FieldType
    if definition.field_type == types.NUMBER:
        try:
            return int(raw)
        except ValueError:
            try:
                return float(raw)
            except ValueError:
                raise ValueError("Enter a number.")
    if definition.field_type == types.BOOLEAN:
        lowered = raw.strip().lower()
        if lowered in {"true", "1", "yes", "y"}:
            return True
        if lowered in {"false", "0", "no", "n"}:
            return False
        raise ValueError("Enter true or false.")
    # date stays an ISO string, choice/text stay strings — validated downstream.
    return raw


def clean_row(project, mapped):
    """Validate one mapped row. Returns ``(data, errors)`` where ``data`` holds
    the cleaned task kwargs (incl. ``custom_fields``) and ``errors`` maps a
    target field to its message. No database writes."""
    errors = {}
    data = {}

    title = mapped.get("title", "").strip()
    if not title:
        errors["title"] = "This field is required."
    else:
        data["title"] = title

    if "notes" in mapped:
        data["notes"] = mapped["notes"]

    if "due_date" in mapped:
        try:
            datetime.date.fromisoformat(mapped["due_date"])
            data["due_date"] = mapped["due_date"]
        except ValueError:
            errors["due_date"] = "Enter a date (YYYY-MM-DD)."

    project_type = project.project_type if project.project_type_id else None

    # Custom fields: coerce strings to types, then validate against the schema.
    definitions = {}
    if project_type is not None:
        definitions = {d.name: d for d in project_type.field_definitions.all()}
    raw_custom = {}
    for target, value in mapped.items():
        if not target.startswith(CUSTOM_PREFIX):
            continue
        name = target[len(CUSTOM_PREFIX):]
        definition = definitions.get(name)
        if definition is None:
            errors[target] = "Unknown field."
            continue
        try:
            raw_custom[name] = _coerce_custom(definition, value)
        except ValueError as exc:
            errors[target] = str(exc)
    try:
        cleaned_custom = validate_custom_fields(project_type, raw_custom)
        data["custom_fields"] = cleaned_custom
    except ValidationError as exc:
        for name, messages in exc.message_dict.items():
            key = f"{CUSTOM_PREFIX}{name}"
            # Don't clobber a more specific coercion error already recorded.
            errors.setdefault(
                key, messages[0] if isinstance(messages, list) else messages
            )

    # Status against the workflow (empty resolves to the default / "").
    try:
        data["status"] = validate_status(project_type, mapped.get("status", ""))
    except ValidationError as exc:
        errors["status"] = exc.messages[0]

    return data, errors


def preview(project, rows, mapping):
    """Validate every row without writing. Returns a list of result dicts with
    ``row_number`` (1-based, header excluded), the mapped ``values``, cleaned
    ``data``, ``errors``, and ``valid``."""
    results = []
    for index, row in enumerate(rows, start=1):
        mapped = map_row(row, mapping)
        data, errors = clean_row(project, mapped)
        results.append(
            {
                "row_number": index,
                "values": mapped,
                "data": data,
                "errors": errors,
                "valid": not errors,
            }
        )
    return results


def commit(project, rows, mapping, owner):
    """Create tasks for the valid rows; skip invalid ones and report them.

    Returns ``{"created": int, "skipped": [{"row_number", "errors"}]}``.
    """
    created = 0
    skipped = []
    for result in preview(project, rows, mapping):
        if not result["valid"]:
            skipped.append(
                {"row_number": result["row_number"], "errors": result["errors"]}
            )
            continue
        data = result["data"]
        task = Task(
            project=project,
            owner=owner,
            title=data["title"],
            notes=data.get("notes", ""),
            due_date=data.get("due_date") or None,
            status=data.get("status", ""),
            custom_fields=data.get("custom_fields", {}),
        )
        task.save()  # keeps is_done in sync with the workflow status
        log_created(task, owner)
        created += 1
    return {"created": created, "skipped": skipped}
